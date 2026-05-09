import os
import re
import io
from datetime import datetime

from flask import Flask, render_template, request, redirect, url_for, send_from_directory, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash

from PyPDF2 import PdfReader
from rapidfuzz import fuzz
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


BASE_DIR = os.path.abspath(os.path.dirname(__file__))

app = Flask(__name__)
app.config["SECRET_KEY"] = "cambia-esta-clave-en-produccion"
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///" + os.path.join(BASE_DIR, "instance", "database.db")
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["UPLOAD_FOLDER"] = os.path.join(BASE_DIR, "uploads")
app.config["MAX_CONTENT_LENGTH"] = 40 * 1024 * 1024

ALLOWED_EXTENSIONS = {".pdf"}

os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
os.makedirs(os.path.join(BASE_DIR, "instance"), exist_ok=True)

db = SQLAlchemy(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"


# ============================================================
# MODELOS
# ============================================================

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    cvs = db.relationship("CV", backref="owner", lazy=True)


class CV(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(255), nullable=False)
    original_name = db.Column(db.String(255), nullable=False)
    extracted_text = db.Column(db.Text, default="")
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)


class Analysis(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    job_description = db.Column(db.Text, nullable=False)
    best_candidate = db.Column(db.String(255))
    best_score = db.Column(db.Float, default=0)
    summary = db.Column(db.Text)
    user_id = db.Column(db.Integer, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


# ============================================================
# MOTOR ATS INTELIGENTE LOCAL
# ============================================================

STOPWORDS = {
    "y", "o", "de", "del", "la", "el", "los", "las", "un", "una", "unos", "unas",
    "que", "con", "para", "por", "en", "al", "lo", "su", "sus", "como", "se",
    "es", "son", "ser", "tener", "tenga", "tiene", "requiere", "requiero",
    "busco", "necesito", "candidato", "persona", "perfil", "buen", "buena",
    "muy", "tambien", "también", "este", "esta", "estos", "estas", "cargo",
    "puesto", "empresa", "trabajo", "trabajar", "alguien", "ideal"
}

TECH_SKILLS = {
    "redes": ["redes", "networking", "tcp/ip", "lan", "wan", "vlan", "subnetting", "routing", "switching", "router", "switch"],
    "cisco": ["cisco", "packet tracer", "ccna", "ccnp", "ios cisco"],
    "fibra optica": ["fibra optica", "fibra óptica", "ftth", "fusion fibra", "empalme fibra", "otdr", "vfl"],
    "soporte tecnico": ["soporte tecnico", "soporte técnico", "helpdesk", "mesa de ayuda", "atencion usuarios", "atención usuarios"],
    "linux": ["linux", "ubuntu", "debian", "raspberry", "bash", "shell"],
    "windows": ["windows", "active directory", "directorio activo", "office 365", "microsoft 365"],
    "programacion": ["python", "javascript", "html", "css", "flask", "sql", "api", "backend", "frontend"],
    "cloud": ["aws", "azure", "gcp", "nube", "cloud", "ec2", "s3", "iam", "vpc", "lambda"],
    "base de datos": ["sql", "sqlite", "mysql", "mariadb", "postgresql", "base de datos"],
    "ciberseguridad": ["seguridad", "ciberseguridad", "firewall", "vpn", "hardening", "ids", "ips"],
    "excel": ["excel", "excel avanzado", "tablas dinamicas", "tablas dinámicas", "power bi"],
    "rrhh": ["recursos humanos", "rrhh", "reclutamiento", "seleccion", "selección", "entrevista"],
    "ventas": ["ventas", "vendedor", "comercial", "ejecutivo comercial", "sales"],
    "administracion": ["administracion", "administración", "gestion", "gestión", "coordinacion", "coordinación"]
}

CERTIFICATIONS = [
    "ccna", "ccnp", "aws", "aws certified", "scrum", "itil", "comptia",
    "security+", "network+", "azure fundamentals", "google cloud",
    "diplomado", "certificado", "certificacion", "certificación",
    "titulo", "título", "licencia"
]

LANGUAGE_PATTERNS = {
    "Inglés": ["ingles", "inglés", "english", "bilingue", "bilingüe", "b2", "c1", "c2"],
    "Francés": ["frances", "francés", "french"],
    "Español": ["espanol", "español", "spanish", "castellano"]
}

ROLE_GROUPS = {
    "tecnico redes": ["tecnico en redes", "técnico en redes", "network technician", "soporte redes", "infraestructura redes"],
    "soporte ti": ["soporte ti", "soporte tecnico", "helpdesk", "mesa de ayuda", "analista soporte"],
    "desarrollador": ["desarrollador", "programador", "developer", "software engineer"],
    "administrativo": ["administrativo", "asistente administrativo", "secretaria", "recepcionista"],
    "rrhh": ["reclutador", "analista rrhh", "recursos humanos", "seleccion de personal"],
    "ventas": ["vendedor", "ejecutivo comercial", "asesor comercial", "sales"],
    "contabilidad": ["contador", "contabilidad", "finanzas", "tesoreria", "auditoria"]
}


def normalize_text(text):
    if not text:
        return ""

    text = text.lower()

    replacements = {
        "á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u",
        "ñ": "n", "ü": "u"
    }

    for a, b in replacements.items():
        text = text.replace(a, b)

    text = re.sub(r"[^a-z0-9\s/+#.-]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def extract_pdf_text(path):
    text = ""

    try:
        reader = PdfReader(path)

        for page in reader.pages:
            text += " " + (page.extract_text() or "")

    except Exception:
        text = ""

    return text.strip()


def allowed_file(filename):
    _, ext = os.path.splitext(filename.lower())
    return ext in ALLOWED_EXTENSIONS


def user_upload_folder():
    folder = os.path.join(app.config["UPLOAD_FOLDER"], str(current_user.id))
    os.makedirs(folder, exist_ok=True)
    return folder


def extract_keywords(job_description):
    text = normalize_text(job_description)
    tokens = [w for w in text.split() if len(w) > 2 and w not in STOPWORDS]

    phrases = []

    important_phrases = [
        "fibra optica", "soporte tecnico", "mesa de ayuda", "recursos humanos",
        "trabajo bajo presion", "bajo presion", "resolucion de problemas",
        "gestion de proyectos", "base de datos", "active directory",
        "excel avanzado", "atencion usuarios", "atencion al cliente"
    ]

    for phrase in important_phrases:
        if normalize_text(phrase) in text:
            phrases.append(normalize_text(phrase))

    return sorted(set(tokens + phrases))


def detect_skills(text):
    text_norm = normalize_text(text)
    found = {}

    for category, variants in TECH_SKILLS.items():
        hits = []

        for variant in variants:
            variant_norm = normalize_text(variant)

            if re.search(r"\b" + re.escape(variant_norm) + r"\b", text_norm):
                hits.append(variant)

        if hits:
            found[category] = sorted(set(hits))

    return found


def detect_certifications(text):
    text_norm = normalize_text(text)
    found = []

    for cert in CERTIFICATIONS:
        cert_norm = normalize_text(cert)

        if re.search(r"\b" + re.escape(cert_norm) + r"\b", text_norm):
            found.append(cert.upper() if len(cert) <= 5 else cert.title())

    return sorted(set(found))


def detect_languages(text):
    text_norm = normalize_text(text)
    found = []

    for language, patterns in LANGUAGE_PATTERNS.items():
        for pattern in patterns:
            if re.search(r"\b" + re.escape(normalize_text(pattern)) + r"\b", text_norm):
                found.append(language)
                break

    return sorted(set(found))


def detect_roles(text):
    text_norm = normalize_text(text)
    found = []

    for role, variants in ROLE_GROUPS.items():
        for variant in variants:
            if re.search(r"\b" + re.escape(normalize_text(variant)) + r"\b", text_norm):
                found.append(role)
                break

    return sorted(set(found))


def detect_experience_years(text):
    """
    Detecta experiencia SOLO cuando el CV dice explícitamente:
    'X años de experiencia', 'experiencia de X años', etc.
    Evita confundir edad, fechas, RUT, teléfono o años sueltos.
    """
    text_norm = normalize_text(text)

    # Eliminar frases de edad para no confundirlas.
    text_norm = re.sub(r"\b\d{1,2}\s*anos\s*(de)?\s*edad\b", " ", text_norm)
    text_norm = re.sub(r"\bedad\s*[:\-]?\s*\d{1,2}\b", " ", text_norm)

    patterns = [
        r"\b(\d{1,2})\s*anos\s*(de)?\s*experiencia\b",
        r"\b(\d{1,2})\s*anos\s*(de)?\s*experiencia laboral\b",
        r"\b(\d{1,2})\s*anos\s*(de)?\s*experiencia profesional\b",
        r"\bexperiencia\s*(de)?\s*(\d{1,2})\s*anos\b",
        r"\bmas de\s*(\d{1,2})\s*anos\s*(de)?\s*experiencia\b",
        r"\bminimo\s*(\d{1,2})\s*anos\s*(de)?\s*experiencia\b"
    ]

    detected = []

    for pattern in patterns:
        matches = re.findall(pattern, text_norm)

        for match in matches:
            if isinstance(match, tuple):
                for item in match:
                    if str(item).isdigit():
                        value = int(item)
                        if 0 < value <= 40:
                            detected.append(value)
            else:
                if str(match).isdigit():
                    value = int(match)
                    if 0 < value <= 40:
                        detected.append(value)

    return max(detected) if detected else 0


def score_list_overlap(required_terms, cv_text):
    cv_norm = normalize_text(cv_text)

    exact_matches = []
    fuzzy_matches = []

    for term in required_terms:
        term_norm = normalize_text(term)

        if not term_norm:
            continue

        if re.search(r"\b" + re.escape(term_norm) + r"\b", cv_norm):
            exact_matches.append(term_norm)
        else:
            ratio = fuzz.partial_ratio(term_norm, cv_norm)
            if ratio >= 90:
                fuzzy_matches.append(term_norm)

    total = max(len(required_terms), 1)
    score = ((len(exact_matches) * 1.0) + (len(fuzzy_matches) * 0.65)) / total

    return min(score, 1), sorted(set(exact_matches + fuzzy_matches))


def infer_required_skill_categories(job_description):
    job_norm = normalize_text(job_description)
    required = []

    for category, variants in TECH_SKILLS.items():
        for variant in variants + [category]:
            if normalize_text(variant) in job_norm:
                required.append(category)
                break

    return sorted(set(required))


def infer_required_roles(job_description):
    job_norm = normalize_text(job_description)
    required = []

    for role, variants in ROLE_GROUPS.items():
        for variant in variants + [role]:
            if normalize_text(variant) in job_norm:
                required.append(role)
                break

    return sorted(set(required))


def calculate_candidate_score(job_description, cv_text):
    keywords = extract_keywords(job_description)

    required_skill_categories = infer_required_skill_categories(job_description)
    required_roles = infer_required_roles(job_description)

    detected_skills = detect_skills(cv_text)
    detected_certs = detect_certifications(cv_text)
    detected_languages = detect_languages(cv_text)
    detected_roles = detect_roles(cv_text)
    experience_years = detect_experience_years(cv_text)

    keyword_score, keyword_matches = score_list_overlap(keywords, cv_text)

    if required_skill_categories:
        matched_skill_categories = [
            category for category in required_skill_categories
            if category in detected_skills
        ]
        skill_score = len(matched_skill_categories) / max(len(required_skill_categories), 1)
    else:
        matched_skill_categories = []
        skill_score = min(len(detected_skills) / 4, 1)

    if required_roles:
        matched_roles = [
            role for role in required_roles
            if role in detected_roles
        ]
        role_score = len(matched_roles) / max(len(required_roles), 1)
    else:
        matched_roles = []
        role_score = 0.5 if detected_roles else 0

    job_norm = normalize_text(job_description)

    cert_requested = any(normalize_text(c) in job_norm for c in CERTIFICATIONS)
    cert_score = 1 if detected_certs and cert_requested else 0.5 if detected_certs else 0

    language_requested = any(
        normalize_text(pattern) in job_norm
        for patterns in LANGUAGE_PATTERNS.values()
        for pattern in patterns
    )
    language_score = 1 if detected_languages and language_requested else 0.5 if detected_languages else 0

    experience_requested = "experiencia" in job_norm or "anos" in job_norm or "años" in job_description.lower()
    experience_score = min(experience_years / 5, 1) if experience_years and experience_requested else 0

    final_score = (
        skill_score * 0.35 +
        keyword_score * 0.25 +
        role_score * 0.15 +
        cert_score * 0.10 +
        language_score * 0.05 +
        experience_score * 0.10
    ) * 100

    matches = []

    matches.extend(keyword_matches[:8])
    matches.extend(matched_skill_categories)
    matches.extend(matched_roles)
    matches.extend(detected_certs[:4])
    matches.extend(detected_languages[:3])

    metadata = {
        "keywords": keywords,
        "required_skill_categories": required_skill_categories,
        "detected_skills": detected_skills,
        "detected_certifications": detected_certs,
        "detected_languages": detected_languages,
        "detected_roles": detected_roles,
        "matched_skill_categories": matched_skill_categories,
        "matched_roles": matched_roles,
        "experience_years": experience_years,
        "keyword_score": round(keyword_score * 100, 1),
        "skill_score": round(skill_score * 100, 1),
        "role_score": round(role_score * 100, 1)
    }

    return round(min(final_score, 100), 1), sorted(set(matches))[:25], metadata


def generate_ai_style_summary(candidate_name, score, matches, metadata):
    detected_skills = metadata.get("detected_skills", {})
    detected_certs = metadata.get("detected_certifications", [])
    detected_languages = metadata.get("detected_languages", [])
    detected_roles = metadata.get("detected_roles", [])
    matched_skill_categories = metadata.get("matched_skill_categories", [])
    matched_roles = metadata.get("matched_roles", [])
    experience_years = metadata.get("experience_years", 0)

    summary = f"El perfil {candidate_name} fue priorizado por el motor ATS debido a su nivel de compatibilidad con la descripción ingresada. "

    if matched_roles:
        summary += f"Se detectó alineación con el tipo de cargo solicitado: {', '.join(matched_roles)}. "

    if matched_skill_categories:
        summary += f"Las áreas técnicas más relevantes encontradas fueron: {', '.join(matched_skill_categories)}. "
    elif detected_skills:
        summary += f"El CV presenta señales técnicas en áreas como: {', '.join(list(detected_skills.keys())[:6])}. "
    else:
        summary += "No se detectaron suficientes habilidades técnicas explícitas, por lo que se recomienda revisar manualmente el CV. "

    if detected_certs:
        summary += f"También se identificaron posibles certificaciones o respaldos formativos: {', '.join(detected_certs[:5])}. "

    if detected_languages:
        summary += f"Idiomas detectados: {', '.join(detected_languages)}. "

    if experience_years > 0:
        summary += f"El documento menciona explícitamente alrededor de {experience_years} año(s) de experiencia laboral. "
    else:
        summary += "No se encontró una frase explícita que indique años totales de experiencia laboral, por lo que este punto debe validarse en entrevista. "

    if matches:
        summary += f"Coincidencias relevantes usadas por el ranking: {', '.join(matches[:8])}. "

    if score >= 75:
        summary += "Conclusión: es un candidato altamente alineado y recomendable para avanzar a la siguiente etapa."
    elif score >= 50:
        summary += "Conclusión: es un candidato con ajuste interesante, pero requiere validación humana antes de avanzar."
    elif score >= 30:
        summary += "Conclusión: el perfil tiene coincidencias parciales, por lo que puede considerarse como alternativa secundaria."
    else:
        summary += "Conclusión: la compatibilidad detectada es baja y conviene revisar si el cargo fue descrito con suficiente detalle."

    return summary


# ============================================================
# RUTAS
# ============================================================

@app.route("/")
def home():
    return render_template("home.html")


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        if not username or not password:
            flash("Debes ingresar usuario y contraseña.")
            return redirect(url_for("register"))

        existing = User.query.filter_by(username=username).first()

        if existing:
            flash("Ese usuario ya existe.")
            return redirect(url_for("register"))

        user = User(
            username=username,
            password_hash=generate_password_hash(password)
        )

        db.session.add(user)
        db.session.commit()

        login_user(user)
        return redirect(url_for("dashboard"))

    return render_template("register.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        user = User.query.filter_by(username=username).first()

        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            return redirect(url_for("dashboard"))

        flash("Usuario o contraseña incorrectos.")
        return redirect(url_for("login"))

    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("home"))


@app.route("/dashboard", methods=["GET", "POST"])
@login_required
def dashboard():
    cvs = CV.query.filter_by(user_id=current_user.id).order_by(CV.uploaded_at.desc()).all()

    results = []
    recommended = None
    ai_summary = None
    job_description = ""

    if request.method == "POST":
        job_description = request.form.get("job_description", "").strip()

        for cv in cvs:
            score, matches, metadata = calculate_candidate_score(job_description, cv.extracted_text)

            results.append({
                "id": cv.id,
                "name": cv.original_name,
                "score": score,
                "matches": matches,
                "metadata": metadata
            })

        results.sort(key=lambda item: item["score"], reverse=True)

        if results:
            recommended = results[0]
            ai_summary = generate_ai_style_summary(
                recommended["name"],
                recommended["score"],
                recommended["matches"],
                recommended["metadata"]
            )

            analysis = Analysis(
                job_description=job_description,
                best_candidate=recommended["name"],
                best_score=recommended["score"],
                summary=ai_summary,
                user_id=current_user.id
            )

            db.session.add(analysis)
            db.session.commit()

    total_cvs = len(cvs)
    total_analyses = Analysis.query.filter_by(user_id=current_user.id).count()

    return render_template(
        "dashboard.html",
        cvs=cvs,
        results=results,
        recommended=recommended,
        ai_summary=ai_summary,
        job_description=job_description,
        total_cvs=total_cvs,
        total_analyses=total_analyses
    )


@app.route("/upload", methods=["POST"])
@login_required
def upload():
    files = request.files.getlist("cvs")

    for file in files:
        if not file or file.filename == "":
            continue

        if not allowed_file(file.filename):
            continue

        original_name = file.filename
        safe_name = secure_filename(file.filename)
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        filename = f"{timestamp}_{safe_name}"

        path = os.path.join(user_upload_folder(), filename)
        file.save(path)

        extracted_text = extract_pdf_text(path)

        cv = CV(
            filename=filename,
            original_name=original_name,
            extracted_text=extracted_text,
            user_id=current_user.id
        )

        db.session.add(cv)

    db.session.commit()
    return redirect(url_for("dashboard"))


@app.route("/download/<int:cv_id>")
@login_required
def download(cv_id):
    cv = CV.query.get_or_404(cv_id)

    if cv.user_id != current_user.id:
        return redirect(url_for("dashboard"))

    return send_from_directory(
        user_upload_folder(),
        cv.filename,
        as_attachment=True,
        download_name=cv.original_name
    )


@app.route("/delete/<int:cv_id>")
@login_required
def delete_cv(cv_id):
    cv = CV.query.get_or_404(cv_id)

    if cv.user_id == current_user.id:
        path = os.path.join(user_upload_folder(), cv.filename)

        if os.path.exists(path):
            os.remove(path)

        db.session.delete(cv)
        db.session.commit()

    return redirect(url_for("dashboard"))


@app.route("/delete_all")
@login_required
def delete_all():
    cvs = CV.query.filter_by(user_id=current_user.id).all()

    for cv in cvs:
        path = os.path.join(user_upload_folder(), cv.filename)

        if os.path.exists(path):
            os.remove(path)

        db.session.delete(cv)

    db.session.commit()
    return redirect(url_for("dashboard"))


@app.route("/export_excel")
@login_required
def export_excel():
    cvs = CV.query.filter_by(user_id=current_user.id).all()
    last_analysis = Analysis.query.filter_by(user_id=current_user.id).order_by(Analysis.created_at.desc()).first()

    job_description = last_analysis.job_description if last_analysis else ""

    rows = []

    for cv in cvs:
        score, matches, metadata = calculate_candidate_score(job_description, cv.extracted_text)

        rows.append({
            "candidate": cv.original_name,
            "score": score,
            "matches": ", ".join(matches[:12]),
            "skills": ", ".join(metadata.get("detected_skills", {}).keys()),
            "certifications": ", ".join(metadata.get("detected_certifications", [])),
            "languages": ", ".join(metadata.get("detected_languages", [])),
            "roles": ", ".join(metadata.get("detected_roles", [])),
            "experience": metadata.get("experience_years", 0),
            "uploaded_at": cv.uploaded_at.strftime("%Y-%m-%d %H:%M")
        })

    rows.sort(key=lambda item: item["score"], reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ranking ATS"

    title = "Reporte Profesional ATS - Ranking de Candidatos"
    ws.merge_cells("A1:I1")
    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="0F172A")
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = [
        "Ranking", "Candidato", "Score", "Coincidencias",
        "Skills detectadas", "Certificaciones", "Idiomas",
        "Roles detectados", "Experiencia explícita"
    ]

    ws.append([])
    ws.append(headers)

    for index, row in enumerate(rows, start=1):
        ws.append([
            index,
            row["candidate"],
            row["score"],
            row["matches"],
            row["skills"],
            row["certifications"],
            row["languages"],
            row["roles"],
            row["experience"]
        ])

    header_row = 3
    header_fill = PatternFill("solid", fgColor="1E3A8A")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D1D5DB")

    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for row in ws.iter_rows(min_row=4):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    widths = [12, 34, 12, 45, 40, 34, 22, 30, 22]

    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    if len(rows) > 0:
        table_ref = f"A3:I{len(rows) + 3}"
        table = Table(displayName="TablaRankingATS", ref=table_ref)

        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )

        table.tableStyleInfo = style
        ws.add_table(table)

    ws.freeze_panes = "A4"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="ranking_ats_profesional.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


with app.app_context():
    db.create_all()


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
